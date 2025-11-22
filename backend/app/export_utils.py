import csv
import io
from datetime import datetime
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import Order

# Define all available columns for export
ALL_COLUMNS = [
    'orderid',
    'spectrum_reference',
    'customer_account_number',
    'customer_security_code',
    'job_number',
    'business_name',
    'customer_name',
    'customer_email',
    'customer_address',
    'customer_phone',
    'install_date',
    'install_time',
    'has_internet',
    'has_voice',
    'has_tv',
    'has_sbc',
    'has_mobile',
    'mobile_activated',
    'has_wib',
    'notes'
]

# User-friendly column names
COLUMN_LABELS = {
    'orderid': 'Order ID',
    'spectrum_reference': 'Spectrum Reference',
    'customer_account_number': 'Account Number',
    'customer_security_code': 'Security Code',
    'job_number': 'Job Number',
    'business_name': 'Business Name',
    'customer_name': 'Customer Name',
    'customer_email': 'Customer Email',
    'customer_address': 'Customer Address',
    'customer_phone': 'Customer Phone',
    'install_date': 'Install Date',
    'install_time': 'Install Time',
    'has_internet': 'Internet',
    'has_voice': 'Voice Lines',
    'has_tv': 'TV',
    'has_sbc': 'SBC Lines',
    'has_mobile': 'Mobile Lines',
    'mobile_activated': 'Mobile Activated',
    'has_wib': 'WIB',
    'notes': 'Notes'
}

def format_value(value, column_name: str):
    """Format a value for display in export"""
    if value is None:
        return ''

    # Format booleans
    if isinstance(value, bool):
        return 'Yes' if value else 'No'

    # Format dates
    if column_name == 'install_date':
        return value.strftime('%Y-%m-%d') if hasattr(value, 'strftime') else str(value)

    return str(value)

def get_order_data(order: Order, columns: List[str]) -> dict:
    """Extract order data for specified columns"""
    data = {}
    for col in columns:
        value = getattr(order, col, None)
        data[col] = format_value(value, col)
    return data

def generate_excel(orders: List[Order], columns: Optional[List[str]] = None) -> bytes:
    """Generate Excel file from orders with optional column selection"""
    if columns is None:
        columns = ALL_COLUMNS

    # Validate columns
    columns = [col for col in columns if col in ALL_COLUMNS]

    if not columns:
        columns = ALL_COLUMNS

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for idx, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=idx)
        cell.value = COLUMN_LABELS.get(col, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Write data
    for row_idx, order in enumerate(orders, 2):
        order_data = get_order_data(order, columns)
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = order_data[col]
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    # Auto-adjust column widths
    for idx, col in enumerate(columns, 1):
        column_letter = get_column_letter(idx)
        max_length = len(COLUMN_LABELS.get(col, col))

        # Check data lengths
        for row_idx in range(2, len(orders) + 2):
            cell_value = ws.cell(row=row_idx, column=idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))

        # Set column width with some padding
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def generate_csv(orders: List[Order], columns: Optional[List[str]] = None) -> str:
    """Generate CSV file from orders with optional column selection"""
    if columns is None:
        columns = ALL_COLUMNS

    # Validate columns
    columns = [col for col in columns if col in ALL_COLUMNS]

    if not columns:
        columns = ALL_COLUMNS

    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)

    # Write headers
    headers = [COLUMN_LABELS.get(col, col) for col in columns]
    writer.writerow(headers)

    # Write data
    for order in orders:
        order_data = get_order_data(order, columns)
        row = [order_data[col] for col in columns]
        writer.writerow(row)

    return output.getvalue()

def generate_stats_excel(stats: dict, orders: List[Order]) -> bytes:
    """Generate Excel file with statistics report"""
    wb = Workbook()

    # Stats sheet
    ws_stats = wb.active
    ws_stats.title = "Statistics"

    # Define styles
    title_font = Font(bold=True, size=14, color="1E40AF")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    value_font = Font(size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws_stats['A1'] = 'Order Statistics Report'
    ws_stats['A1'].font = title_font
    ws_stats.merge_cells('A1:B1')

    ws_stats['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws_stats['A2'].font = Font(italic=True, size=10)
    ws_stats.merge_cells('A2:B2')

    # Stats data
    row = 4
    stats_data = [
        ('Total Orders', stats.get('total_orders', 0)),
        ('This Week', stats.get('this_week', 0)),
        ('This Month', stats.get('this_month', 0)),
        ('Pending Installs', stats.get('pending_installs', 0)),
        ('', ''),
        ('Product Statistics', ''),
        ('Internet', stats.get('total_internet', 0)),
        ('TV', stats.get('total_tv', 0)),
        ('Mobile Lines', stats.get('total_mobile', 0)),
        ('Voice Lines', stats.get('total_voice', 0)),
    ]

    for label, value in stats_data:
        if label == '' and value == '':
            row += 1
            continue

        if label == 'Product Statistics':
            cell_label = ws_stats.cell(row=row, column=1)
            cell_label.value = label
            cell_label.font = Font(bold=True, size=12, color="059669")
            ws_stats.merge_cells(f'A{row}:B{row}')
            row += 1
            continue

        cell_label = ws_stats.cell(row=row, column=1)
        cell_label.value = label
        cell_label.font = header_font
        cell_label.fill = header_fill
        cell_label.border = border
        cell_label.alignment = Alignment(horizontal="left", vertical="center")

        cell_value = ws_stats.cell(row=row, column=2)
        cell_value.value = value
        cell_value.font = value_font
        cell_value.border = border
        cell_value.alignment = Alignment(horizontal="right", vertical="center")

        row += 1

    # Set column widths
    ws_stats.column_dimensions['A'].width = 25
    ws_stats.column_dimensions['B'].width = 15

    # Add orders sheet if there are orders
    if orders:
        ws_orders = wb.create_sheet("Order Details")

        # Use simplified columns for stats report
        columns = [
            'orderid',
            'spectrum_reference',
            'customer_name',
            'business_name',
            'install_date',
            'has_internet',
            'has_tv',
            'has_mobile',
            'has_voice'
        ]

        # Write headers
        for idx, col in enumerate(columns, 1):
            cell = ws_orders.cell(row=1, column=idx)
            cell.value = COLUMN_LABELS.get(col, col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Write data
        for row_idx, order in enumerate(orders, 2):
            order_data = get_order_data(order, columns)
            for col_idx, col in enumerate(columns, 1):
                cell = ws_orders.cell(row=row_idx, column=col_idx)
                cell.value = order_data[col]
                cell.border = border
                cell.alignment = Alignment(vertical="center")

        # Auto-adjust column widths
        for idx, col in enumerate(columns, 1):
            column_letter = get_column_letter(idx)
            max_length = len(COLUMN_LABELS.get(col, col))

            for row_idx in range(2, len(orders) + 2):
                cell_value = ws_orders.cell(row=row_idx, column=idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            ws_orders.column_dimensions[column_letter].width = min(max_length + 2, 40)

        # Freeze header row
        ws_orders.freeze_panes = "A2"

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
